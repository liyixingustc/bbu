from django.http import JsonResponse, HttpResponse
from django.shortcuts import render

import os
import pandas as pd
import pytz
from datetime import datetime as dt
from dateutil.parser import parse
from django.db.models import Q
from bbu.settings import TIME_ZONE

from ..WorkWorkers.models.models import *
from ..WorkTasks.models.models import *

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Alignment

EST = pytz.timezone(TIME_ZONE)


class PageManager:
    class PanelManager:
        class ModalManager:
            @staticmethod
            def extend_worker_avail(request, *args, **kwargs):
                worker_id = request.POST.get('resourceId')
                start = request.POST.get('start')
                end = request.POST.get('end')

                start = parse(start).replace(tzinfo=EST)
                end = parse(end).replace(tzinfo=EST)
                worker = Workers.objects.get(id=worker_id)

                avails = WorkerAvailable.objects.filter(Q(time_start__range=[start,end])|
                                                        Q(time_end__range=[start, end])|
                                                        Q(time_start__lte=start,time_end__gte=end),
                                                        name__exact=worker)

                if avails.exists():

                    avails_df = pd.DataFrame.from_records(avails.values())
                    start_list = avails_df['time_start'].tolist()
                    start_list.append(start)
                    end_list = avails_df['time_end'].tolist()
                    end_list.append(end)
                    start_new = min(start_list)
                    end_new = max(end_list)
                    duration_new = end_new - start_new
                    ids = avails_df['id'].tolist()

                    WorkerAvailable.objects.filter(id__in=ids).delete()
                    WorkerAvailable.objects.update_or_create(name=worker,
                                                             date=start_new.date(),
                                                             duration=duration_new,
                                                             time_start=start_new,
                                                             time_end=end_new)
                else:
                    duration=end-start
                    WorkerAvailable.objects.update_or_create(name=worker,
                                                             date=start.date(),
                                                             duration=duration,
                                                             time_start=start,
                                                             time_end=end)

                return JsonResponse({})

        class TableManager:
            @staticmethod
            def create(request, *args, **kwargs):

                data = pd.DataFrame()
                if not data.empty:
                    response = data.to_dict(orient='records')
                    return JsonResponse(response, safe=False)
                else:
                    return JsonResponse({})

        class FormManager:
            @staticmethod
            def submit(request, *args, **kwargs):
                ReportType = request.GET.get('ReportType')
                start = request.GET.get('PeriodStart')
                end = request.GET.get('PeriodEnd')

                if start:
                    start = dt.strptime(start,'%Y-%m-%d').replace(tzinfo=EST)
                if end:
                    end = dt.strptime(end,'%Y-%m-%d').replace(tzinfo=EST)

                worker_scheduled = WorkerScheduled.objects.filter(time_start__gte=start, time_end__lte=end)
                if worker_scheduled.exists():
                    worker_scheduled_df = pd.DataFrame.from_records(worker_scheduled.values('name__name',
                                                                                            'date',
                                                                                            'task_id__work_order',
                                                                                            'task_id__priority',
                                                                                            'task_id__description',
                                                                                            'task_id__estimate_hour'))
                else:
                    worker_scheduled_df = pd.DataFrame()

                worker_scheduled_df.rename_axis({'name__name':'Mechanic',
                                                 'date':'date',
                                                 'task_id__work_order':'Work Order',
                                                 'task_id__priority':'Priority',
                                                 'task_id__description':'Description',
                                                 'task_id__estimate_hour':'EST'},axis=1,inplace=True)
                worker_scheduled_df['EST'] = worker_scheduled_df['EST'].apply(lambda x: x.total_seconds()/3600)
                worker_scheduled_df = worker_scheduled_df[['date', 'Mechanic', 'Work Order','Priority', 'Description', 'EST']]

                wb = Workbook()
                ws = wb.active

                date_unique = worker_scheduled_df.date.unique()

                start_row = 2
                start_col = 1
                start_row_data = start_row + 1
                start_col_data = start_col
                length = 5

                for date in date_unique:

                    ws.merge_cells(start_row=start_row,
                                   start_column=start_col,
                                   end_row=start_row,
                                   end_column=start_col+length-1)
                    # date row
                    date_cell = ws.cell(row=start_row, column=start_col, value=date)
                    date_cell.alignment = Alignment(horizontal="center", vertical="center")

                    df = worker_scheduled_df[worker_scheduled_df['date'] == date]
                    df = df.drop('date',axis=1)

                    for r in dataframe_to_rows(df, index=False, header=True):
                        for i in r:
                            ws.cell(row=start_row_data, column=start_col_data, value=i)
                            start_col_data += 1

                        start_row_data += 1
                        start_col_data = start_col

                    # reset row and column
                    start_row_data = start_row + 1
                    start_col += length + 1
                    start_col_data = start_col

                # setting title
                ws.merge_cells(start_row=1,
                               start_column=1,
                               end_row=1,
                               end_column=start_col - 2)
                title_cell = ws.cell(row=1,column=1,value='Bread Week Down Day Work Assignments')
                title_cell.alignment = Alignment(horizontal="center", vertical="center")


                wb.save('WorkSchedule/WorkReports/result.xlsx')
                wb.close()


                return JsonResponse({})

            @staticmethod
            def download(request, *args, **kwargs):

                # try:
                output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'result.xlsx')
                file_path = open(output_path, 'rb').read()
                response = HttpResponse(file_path, content_type='xlsx')
                response['Content-Disposition'] = "attachment; filename='result.xlsx'"
                # except Exception as e:
                #     response = JsonResponse({})

                return response
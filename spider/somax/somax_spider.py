from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pyvirtualdisplay import Display

from bs4 import BeautifulSoup as bs
import os
import sys
import shutil
import glob
import time

import numpy as np
import pandas as pd
from openpyxl import load_workbook

sys.path.append('../..')
import django
from bbu.settings import BASE_DIR, MEDIA_ROOT

os.environ['DJANGO_SETTINGS_MODULE'] = 'bbu.settings'
django.setup()

from WorkSchedule.WorkConfig.processor.EquipmentLoadProcessor import EquipmentLoadProcessor
from WorkSchedule.WorkConfig.processor.PMsLoadProcessor import PMsLoadProcessor
from WorkSchedule.WorkConfig.processor.TasksLoadProcessor import TasksLoadProcessor

from WorkSchedule.WorkConfig.models.models import *
from WorkSchedule.WorkWorkers.models.models import *
from WorkSchedule.WorkTasks.models.models import *

from DAO.WorkScheduleDataDAO import WorkScheduleDataDAO
from DAO.WorkScheduleReviseDAO import WorkScheduleReviseDAO


class SomaxSpider:

    DISPLAY = False

    account = 'BBUGRNATU'
    password = 'ARTHUR'

    somax_dashboard_url = 'https://somaxonline.somax.com/DashboardMain.aspx'
    somax_login_url = 'https://somaxonline.somax.com/Login.aspx'
    somax_equipment_url = 'https://somaxonline.somax.com/EquipmentSearch.aspx'
    somax_equipment_edit_url = 'https://somaxonline.somax.com/EquipmentEdit.aspx'
    somax_task_url = 'https://somaxonline.somax.com/WorkOrderSearch.aspx'
    somax_task_edit_url = 'https://somaxonline.somax.com/WorkOrderEdit.aspx'
    somax_pm_url = 'https://somaxonline.somax.com/PreventiveMaintenanceSearch.aspx'
    somax_pm_edit_url = 'https://somaxonline.somax.com/PreventiveMaintenanceDetails.aspx'

    somax_logo_id = 'imgLogo'
    somax_page_title_id = 'ctlPageDetails_lblPageTitle'
    somax_table_export_div_id = 'MainContent_uicSearchHeader_dxBtnExport'
    somax_task_filter_work_order_id = 'MainContent_grdResults_DXFREditorcol0_I'
    somax_task_first_row_work_order_a_id = 'MainContent_grdResults_cell0_0_ColClientLookUpId_0'
    somax_task_first_row_work_order_td_id = 'MainContent_grdResults_tccell0_0'
    somax_task_detail_actual_open_span_id = 'MainContent_cpcActuals_lblHeaderCollapsed'
    somax_task_detail_actual_tab_a_id = 'MainContent_cpcActuals_ctl00_ASPxPageControlActuals_T1'
    somax_task_detail_actual_table_id = 'MainContent_cpcActuals_ctl00_ASPxPageControlActuals_grdWorkOrderLabor_DXMainTable'

    somax_label_scheduling_url = 'https://somaxonline.somax.com/UpdatedLaborScheduling.aspx'
    somax_label_scheduling_table_id = 'MainContent_ASPxCallbackPanel1_gridScheduled_DXMainTable'
    somax_label_scheduling_table_first_row_id = 'MainContent_ASPxCallbackPanel1_gridScheduled_DXDataRow0'
    somax_label_scheduling_table_edit_input_id = 'MainContent_ASPxCallbackPanel1_gridScheduled_DXEditor4_I'
    somax_label_scheduling_date_input_id = 'MainContent_uicSchuduledDate_datetimeControl_I'
    somax_label_scheduling_assigned_input_id = 'MainContent_ddlPersonnel_ddldxControl_I'
    somax_label_scheduling_add_input_id = 'MainContent_ASPxCallbackPanel1_btnAvailableWork'
    somax_label_scheduling_loading_id = 'MainContent_ASPxCallbackPanel1_LPV'
    somax_label_scheduling_loading_grey_table_id = 'MainContent_ASPxCallbackPanel1_gridScheduled_LPV'
    somax_label_scheduling_loading_grey_span_id = 'MainContent_ASPxCallbackPanel1_gridScheduled_TL'
    somax_label_scheduling_modal_include_all_input_id = 'MainContent_ASPxCallbackPanel1_pcLogin_Panel1_chk_Order'
    somax_label_scheduling_modal_select_all_input_id = 'MainContent_ASPxCallbackPanel1_pcLogin_Panel1_gridAvailable_DXSelAllBtn0_D'
    somax_label_scheduling_modal_select_all_input_class_unchecked = 'dxWeb_edtCheckBoxUnchecked_GridTheme dxICheckBox_GridTheme dxichSys'
    somax_label_scheduling_modal_select_all_input_class_checked = 'dxWeb_edtCheckBoxChecked_GridTheme dxICheckBox_GridTheme dxichSys'
    somax_label_scheduling_modal_select_all_input_class_partial_checked = 'dxWeb_edtCheckBoxGrayed_GridTheme dxICheckBox_GridTheme dxichSys'
    somax_label_scheduling_modal_filter_work_order_id = 'MainContent_ASPxCallbackPanel1_pcLogin_Panel1_gridAvailable_DXFREditorcol1_I'
    somax_label_scheduling_modal_table = 'MainContent_ASPxCallbackPanel1_pcLogin_Panel1_gridAvailable_DXMainTable'
    somax_label_scheduling_modal_first_row_id = 'MainContent_ASPxCallbackPanel1_pcLogin_Panel1_gridAvailable_DXDataRow0'
    somax_label_scheduling_modal_first_row_check_id = 'MainContent_ASPxCallbackPanel1_pcLogin_Panel1_gridAvailable_DXSelBtn0'
    somax_label_scheduling_modal_first_row_empty_id = 'MainContent_ASPxCallbackPanel1_pcLogin_Panel1_gridAvailable_DXEmptyRow'
    somax_label_scheduling_modal_loading_id = 'MainContent_ASPxCallbackPanel1_pcLogin_Panel1_gridAvailable_TL'
    somax_label_scheduling_modal_loading_grey_table_id = 'MainContent_ASPxCallbackPanel1_pcLogin_Panel1_gridAvailable_LPV'
    somax_label_scheduling_modal_loading_grey_span_id = 'MainContent_ASPxCallbackPanel1_pcLogin_Panel1_gridAvailable_TL'
    somax_label_scheduling_modal_add_input_id = 'MainContent_ASPxCallbackPanel1_pcLogin_Panel1_btnAdd'
    somax_label_scheduling_modal_close_id = 'MainContent_ASPxCallbackPanel1_pcLogin_Panel1_dxImgPopUpClose'
    somax_label_scheduling_modal_repeat_id = 'MainContent_ASPxCallbackPanel1_PopupMessage_PW-1'
    somax_label_scheduling_modal_repeat_ok_id = 'MainContent_ASPxCallbackPanel1_PopupMessage_btnOk'

    download_path = os.path.join(BASE_DIR, MEDIA_ROOT, 'spider', 'somax')

    def __init__(self, account=None, password=None):
        self.init_folder()
        if not self.DISPLAY:
            self.display = Display(visible=0, size=(1440, 900))
            self.display.start()
        self.driver = self.chromedriver()
        # self.login(account, password)
        # self.cookies = self.driver.get_cookies()

    @staticmethod
    def init_folder():
        equipment_folder_path = os.path.join(BASE_DIR, MEDIA_ROOT, 'spider/somax/equipment')
        pm_folder_path = os.path.join(BASE_DIR, MEDIA_ROOT, 'spider/somax/pm')
        task_folder_path = os.path.join(BASE_DIR, MEDIA_ROOT, 'spider/somax/task')

        os.makedirs(equipment_folder_path, exist_ok=True)
        os.makedirs(pm_folder_path, exist_ok=True)
        os.makedirs(task_folder_path, exist_ok=True)

    def firefoxdriver(self):
        driver = webdriver.Firefox(executable_path='/usr/bin/geckodriver')

        return driver

    def chromedriver(self):
        options = webdriver.ChromeOptions()
        prefs = {'profile.default_content_settings.popups': 0, 'download.default_directory': self.download_path}
        options.add_experimental_option('prefs', prefs)
        driver = webdriver.Chrome(chrome_options=options,executable_path='/usr/bin/chromedriver')
        return driver

    def phantomjsdriver(self):

        driver = webdriver.PhantomJS()

        return driver

    def login(self, account=None, password=None):
        if account and password:
            self.account = account
            self.password = password
        self.driver.get(self.somax_login_url)
        self.driver.find_element_by_id('txtUserName').clear()
        self.driver.find_element_by_id('txtUserName').send_keys(self.account)
        self.driver.find_element_by_id('txtPassword').clear()
        self.driver.find_element_by_id('txtPassword').send_keys(self.password)
        self.driver.find_element_by_id('btnLogin').click()

        while True:
            if self.driver.current_url != self.somax_login_url:
                break
            time.sleep(1)
        return True

    def equipment_spider(self):
        try:
            self.driver.get(self.somax_equipment_url)
            current_url = self.driver.current_url
            if current_url == self.somax_login_url:
                self.login()
            before = os.listdir(self.download_path)
            self.get_and_ready(self.somax_equipment_url)
            element_export = WebDriverWait(self.driver, 60).until(
                EC.element_to_be_clickable((By.ID, self.somax_table_export_div_id)))
            self.driver.execute_script("arguments[0].click();", element_export)
            self.download_finished(file_name='EquipmentSearch.csv')
            after = os.listdir(self.download_path)
            self.driver.execute_script("window.stop();")
            filename = self.get_download_file_name(before, after, self.download_path)
            if not filename:
                return None
            else:
                target_path = os.path.join(self.download_path, 'equipment')
                r = glob.glob(os.path.join(target_path, '*'))
                for i in r:
                        os.remove(i)

                shutil.move(os.path.join(self.download_path, filename),
                            os.path.join(target_path, filename))

            self.driver.quit()
            if not self.DISPLAY:
                self.display.stop()
            file_path = os.path.join(target_path, filename)
            EquipmentLoadProcessor.equipment_load_processor([file_path])
        except Exception as e:
            self.driver.quit()
            print(e)

        return True

    def pm_spider(self):
        try:
            self.driver.get(self.somax_pm_url)
            current_url = self.driver.current_url
            if current_url == self.somax_login_url:
                self.login()

            before = os.listdir(self.download_path)
            self.get_and_ready(self.somax_pm_url)
            element_export = WebDriverWait(self.driver, 60).until(
                EC.element_to_be_clickable((By.ID, self.somax_table_export_div_id)))
            self.driver.execute_script("arguments[0].click();", element_export)
            self.download_finished(file_name='PreventiveMaintenance.csv')
            after = os.listdir(self.download_path)
            self.driver.execute_script("window.stop();")
            filename = self.get_download_file_name(before, after, self.download_path)
            if not filename:
                return None
            else:
                target_path = os.path.join(self.download_path, 'pm')
                r = glob.glob(os.path.join(target_path, '*'))
                for i in r:
                        os.remove(i)

                shutil.move(os.path.join(self.download_path, filename),
                            os.path.join(target_path, filename))

            self.driver.quit()
            if not self.DISPLAY:
                self.display.stop()

            file_path = os.path.join(target_path, filename)
            PMsLoadProcessor.pms_load_processor([file_path])
        except Exception as e:
            self.driver.quit()
            print(e)
        return True

    def task_spider(self):
        try:
            self.driver.get(self.somax_task_url)
            current_url = self.driver.current_url
            if current_url == self.somax_login_url:
                self.login()

            before = os.listdir(self.download_path)
            self.get_and_ready(self.somax_task_url)
            element_export = WebDriverWait(self.driver, 60).until(
                EC.element_to_be_clickable((By.ID, self.somax_table_export_div_id)))
            self.driver.execute_script("arguments[0].click();", element_export)
            self.download_finished(file_name='WorkOrderSearch.csv')
            after = os.listdir(self.download_path)
            self.driver.execute_script("window.stop();")
            filename = self.get_download_file_name(before, after, self.download_path)
            if not filename:
                return None
            else:
                target_path = os.path.join(self.download_path, 'task')
                r = glob.glob(os.path.join(target_path, '*'))
                for i in r:
                        os.remove(i)

                shutil.move(os.path.join(self.download_path, filename),
                            os.path.join(target_path, filename))

            self.driver.quit()
            if not self.DISPLAY:
                self.display.stop()

            file_path = os.path.join(target_path, filename)
            TasksLoadProcessor.tasks_load_processor([file_path])
        except Exception as e:
            self.driver.quit()
            print(e)

        return True

    def task_edit_spider(self, work_orders_dict=None):
        if not work_orders_dict:
            return True

        # filter out no change record
        worker_order_df = pd.DataFrame.from_records(work_orders_dict)
        # worker_order_df = worker_order_df[worker_order_df['current_status'] != worker_order_df['current_status_somax']]
        # if worker_order_df.empty:
        #     return True

        # wait and check again
        # time.sleep(100)

        for index, row in worker_order_df.iterrows():
            work_order = row['work_order']
            tasks_obj = Tasks.objects.filter(work_order__exact=work_order)
            if tasks_obj.exists():
                current_status = tasks_obj[0].current_status
            else:
                continue

            worker_order_df.set_value(index, 'current_status', current_status)

        # filter out no change record
        # worker_order_df = worker_order_df[worker_order_df['current_status'] != worker_order_df['current_status_somax']]
        # if worker_order_df.empty:
        #     return True

        # edit somax
        self.driver.get(self.somax_task_url)
        current_url = self.driver.current_url
        if current_url == self.somax_login_url:
            self.login()

        self.get_and_ready(self.somax_task_url)

        for work_order_dict in work_orders_dict:

            work_order = work_order_dict['work_order']
            current_status = work_order_dict['current_status']

            self.get_work_order_detail(work_order)

        self.driver.quit()
        if not self.DISPLAY:
            self.display.stop()

        return True

    def task_actual_hour_spider(self, start, end):

        # remove
        data = pd.read_excel('Work hour result.xlsx')
        data = data.loc[int(start): int(end)]
        data_unfilled = data[data['hours'].isnull()]
        # data_unfilled = data[data['hours']==-1]

        data['Scheduled'] = data['Scheduled'].apply(lambda x: x.date())
        data['Created'] = data['Created'].apply(lambda x: x.date())

        self.driver.get(self.somax_task_url)
        current_url = self.driver.current_url
        if current_url == self.somax_login_url:
            self.login()

        self.get_and_ready(self.somax_task_url)

        i = 0
        for index, row in data_unfilled.iterrows():
            try:
                work_order = str(row['Work Order'])
                self.get_work_order_detail(work_order)

                element_click_plus = WebDriverWait(self.driver, 60).until(
                    EC.element_to_be_clickable((By.ID, self.somax_task_detail_actual_open_span_id)))
                element_click_plus.click()

                element_click_tab = WebDriverWait(self.driver, 60).until(
                    EC.element_to_be_clickable((By.ID, self.somax_task_detail_actual_tab_a_id)))
                element_click_tab.click()

                element_table = WebDriverWait(self.driver, 60).until(
                    EC.presence_of_element_located((By.ID, self.somax_task_detail_actual_table_id)))

                table_html = element_table.get_attribute('outerHTML')
                soup = bs(table_html, 'lxml')

            # with open("output1.html", "w") as file:
            #     file.write(str(soup))
            # print(table_html)
            # print(pd.read_html(table_html))

                columns = ['username_somax', 'full_name', 'date', 'hours', 'value']
                table_data = pd.DataFrame(columns=columns)

                for script in soup.find_all('script'):
                    script.extract()
                tr_s = soup.find_all('tr', class_='dxgvDataRow_GridTheme')

                for tr in tr_s:
                    td_s = tr.find_all('td')[1:-1]
                    row = {'username_somax': td_s[0].get_text(),
                           'full_name': td_s[1].get_text(),
                           'date': td_s[2].get_text(),
                           'hours': float(td_s[3].get_text()),
                           'value': float(td_s[4].get_text())
                           }

                    table_data = table_data.append(pd.Series(row), ignore_index=True)

                total_actual = table_data['hours'].sum()

                # data.set_value(index, 'hours', total_actual)

            except Exception as e:
                print(e)
                total_actual = -1
                # data.set_value(index, 'hours', -1)

            self.to_excel(index + 2, total_actual)
            print(i)
            i = i + 1
            self.get_and_ready(self.somax_task_url)

        self.driver.quit()
        if not self.DISPLAY:
            self.display.stop()

        return True

    @classmethod
    def get_schedules_to_somax_spider(cls):
        tasks_records = WorkScheduleDataDAO.get_all_schedules_sync_to_somax()

        if tasks_records.empty:
            return pd.DataFrame()

        tasks_records.sort_values(['worker', 'date'], inplace=True)
        tasks_records['group'] = 0

        group = 0
        worker_last = None
        date_last = None
        for index, row in tasks_records.iterrows():
            worker = row['worker']
            date = row['date']
            if worker_last and date_last:
                if worker == worker_last and date ==date_last:
                    tasks_records.set_value(index, ['group'], group)
                else:
                    group += 1
                    worker_last = worker
                    date_last = date
                    tasks_records.set_value(index, ['group'], group)
            else:
                worker_last = worker
                date_last = date
                tasks_records.set_value(index, ['group'], group)


        # print(tasks_records)

        return tasks_records

    def sync_schedules_to_somax_spider(self):

        tasks_records = self.get_schedules_to_somax_spider()
        if tasks_records.empty:
            return None

        # dummy data
        # date = '2017/12/10'
        # assigned = 'BBUGRNATU'
        # work_orders = ['17001030', '17003808']

        self.driver.get(self.somax_label_scheduling_url)
        current_url = self.driver.current_url
        if current_url == self.somax_login_url:
            self.login()

        self.get_and_ready(self.somax_label_scheduling_url)

        groups = tasks_records['group'].unique()
        for group in groups:
            tasks_group = tasks_records[tasks_records['group'] == group]
            self.sync_schedule_to_somax_spider(tasks_group)
            # self.sync_schedule_to_somax_spider('BBUGRNATU', '2017/12/10', ['17001030', '17003808'])

        self.driver.quit()

        return True

    def sync_schedule_to_somax_spider(self, tasks_group):

        assigned = tasks_group['worker'].unique()[0]
        date = tasks_group['date'].unique()[0]
        date = '{dt.year}/{dt.month}/{dt.day}'.format(dt=date)
        work_orders = tasks_group['work_order'].tolist()
        print('step1')
        # revise assigned
        element_assigned = WebDriverWait(self.driver, 60).until(
            EC.presence_of_element_located((By.ID, self.somax_label_scheduling_assigned_input_id)))
        element_assigned.clear()
        element_assigned.send_keys(assigned)
        element_assigned.send_keys(Keys.ENTER)
        # self.driver.execute_script("arguments[0].setAttribute('value', '{assigned}')".format(assigned=assigned),
        #                            element_assigned)
        self.wait_loading(self.somax_label_scheduling_loading_id)

        print('step2')
        # revise date
        remove_date_readonly_script = 'document.getElementById("{id}").removeAttribute("readonly")'\
            .format(id=self.somax_label_scheduling_date_input_id)
        self.driver.execute_script(remove_date_readonly_script)

        element_date = WebDriverWait(self.driver, 60).until(
            EC.presence_of_element_located((By.ID, self.somax_label_scheduling_date_input_id)))

        date_element = self.driver.find_element_by_id(self.somax_label_scheduling_date_input_id)
        print('abc'+ date_element.get_attribute('innerHTML'))
        element_date.clear()
        element_date.send_keys(date)
        element_date.send_keys(Keys.ENTER)
        print('abc'+ element_date.get_attribute('innerHTML'))
        # self.driver.execute_script("arguments[0].setAttribute('value', '{date}')".format(date=date),
        #                            element_date)
        self.wait_loading(self.somax_label_scheduling_loading_id)
        print('abc'+ element_date.get_attribute('innerHTML'))
        print('step3')
        # click available modal
        element_add = WebDriverWait(self.driver, 60).until(
            EC.presence_of_element_located((By.ID, self.somax_label_scheduling_add_input_id)))
        self.driver.execute_script("arguments[0].click();", element_add)
        self.wait_loading(self.somax_label_scheduling_loading_id)
        print('step4')
        # clear filter
        element_modal_filter_work_order = WebDriverWait(self.driver, 60).until(
            EC.presence_of_element_located((By.ID, self.somax_label_scheduling_modal_filter_work_order_id)))
        if element_modal_filter_work_order.text:
            self.driver.execute_script("arguments[0].setAttribute('value', '')",
                                       element_modal_filter_work_order)
            element_modal_filter_work_order.send_keys(Keys.ENTER)
        print('step5')
        # click include all
        element_modal_include_all = WebDriverWait(self.driver, 60).until(
            EC.presence_of_element_located((By.ID, self.somax_label_scheduling_modal_include_all_input_id)))
        element_modal_include_all_is_selected = element_modal_include_all.is_selected()
        if not element_modal_include_all_is_selected:
            self.driver.execute_script("arguments[0].click();", element_modal_include_all)
            self.wait_loading(self.somax_label_scheduling_loading_id)
        print('step6')
        # clear select all
        element_modal_select_all = WebDriverWait(self.driver, 60).until(
            EC.presence_of_element_located((By.ID, self.somax_label_scheduling_modal_select_all_input_id)))
        element_modal_select_all_class = element_modal_select_all.get_attribute('class')
        if element_modal_select_all_class == self.somax_label_scheduling_modal_select_all_input_class_checked:
            self.driver.execute_script("arguments[0].click();", element_modal_select_all)
        elif element_modal_select_all_class == self.somax_label_scheduling_modal_select_all_input_class_partial_checked:
            self.driver.execute_script("arguments[0].click();", element_modal_select_all)
            self.driver.execute_script("arguments[0].click();", element_modal_select_all)
        print('step7')
        # select work orders
        # work_order = work_orders[0]
        for work_order in work_orders:

            # click include all
            element_modal_include_all = WebDriverWait(self.driver, 60).until(
                EC.presence_of_element_located((By.ID, self.somax_label_scheduling_modal_include_all_input_id)))
            element_modal_include_all_is_selected = element_modal_include_all.is_selected()
            if not element_modal_include_all_is_selected:
                self.driver.execute_script("arguments[0].click();", element_modal_include_all)
                self.wait_loading(self.somax_label_scheduling_loading_id)

            # select work_order
            element_modal_filter_work_order = WebDriverWait(self.driver, 60).until(
                EC.presence_of_element_located((By.ID, self.somax_label_scheduling_modal_filter_work_order_id)))
            self.driver.execute_script("arguments[0].setAttribute('value', '{work_order}')".format(work_order=work_order),
                                       element_modal_filter_work_order)
            element_modal_filter_work_order.send_keys(Keys.ENTER)

            self.wait_loading(self.somax_label_scheduling_modal_loading_grey_table_id, timeout=5)

            element_modal_last_row_first_cell_xpath = "//table[@id='{table_id}']/tbody/tr[last()]/td[2]"\
                .format(table_id=self.somax_label_scheduling_modal_table)
            WebDriverWait(self.driver, 20).until(
                EC.text_to_be_present_in_element((By.XPATH, element_modal_last_row_first_cell_xpath), str(work_order)))

            element_modal_last_row_first_cell_element = self.driver.find_elements_by_xpath(element_modal_last_row_first_cell_xpath)
            print(element_modal_last_row_first_cell_element)
            element_modal_last_row_first_cell_element = element_modal_last_row_first_cell_element[0]
            print(element_modal_last_row_first_cell_element.get_attribute('innerHTML'))

            element_modal_first_row_check = WebDriverWait(self.driver, 60).until(
                EC.presence_of_element_located((By.ID, self.somax_label_scheduling_modal_first_row_check_id)))
            self.driver.execute_script("arguments[0].click();", element_modal_first_row_check)
        print('step8')
        # click modal add
        element_modal_add = WebDriverWait(self.driver, 60).until(
            EC.presence_of_element_located((By.ID, self.somax_label_scheduling_modal_add_input_id)))
        # self.driver.execute_script("arguments[0].click();", element_modal_add)
        element_modal_add.click()
        self.wait_loading(self.somax_label_scheduling_loading_id)
        print('step9')
        # click modal close
        element_modal_close = WebDriverWait(self.driver, 60).until(
            EC.presence_of_element_located((By.ID, self.somax_label_scheduling_modal_close_id)))
        self.driver.execute_script("arguments[0].click();", element_modal_close)
        print('step10')
        self.wait_loading(self.somax_label_scheduling_loading_grey_table_id)

        # close repeat window
        element_modal_repeat = self.driver.find_element_by_id(self.somax_label_scheduling_modal_repeat_id)
        if element_modal_repeat:
            element_modal_close = WebDriverWait(self.driver, 60).until(
                EC.presence_of_element_located((By.ID, self.somax_label_scheduling_modal_repeat_ok_id)))
            self.driver.execute_script("arguments[0].click();", element_modal_close)

        # sync schedule hours
        print('step11')
        for index, row in tasks_group.iterrows():
            work_order = row['work_order']
            hours = row['schedule_hour']
            self.sync_schedule_hour_to_somax_spider(work_order, hours)

        time.sleep(3)

        return True

    def sync_schedule_hour_to_somax_spider(self, work_order=None, hours=None):

        if hours:
            hours = str(round(hours, 2))
        else:
            hours = '0'

        # work_order = work_orders[0]
        element_table_first_row = WebDriverWait(self.driver, 60).until(
            EC.presence_of_element_located((By.ID, self.somax_label_scheduling_table_first_row_id)))
        element_table_first_row = self.driver.find_elements_by_id(self.somax_label_scheduling_table_first_row_id)
        if not element_table_first_row:
            return False

        element_table_rows_xpath = "//table[@id='{table_id}']/tbody/tr" \
            .format(table_id=self.somax_label_scheduling_table_id)
        element_table_rows_counts = len(self.driver.find_elements_by_xpath(element_table_rows_xpath)[3:-1])

        for row_num in range(1, element_table_rows_counts+1):

            element_table_row = self.driver.find_elements_by_xpath(element_table_rows_xpath)[row_num+2]
            table_work_order = element_table_row.find_elements_by_tag_name('td')[1].text

            if table_work_order != str(work_order):
                continue

            element_table_hour = element_table_row.find_elements_by_tag_name('td')[4]
            # hours_old = element_table_hour.text
            # if abs(float(hours_old) - float(hours)) < 0.01:
            #     print('abc')
            #     continue

            element_table_hour.click()
            actionChains = ActionChains(self.driver)
            actionChains.double_click(element_table_hour).perform()
            element_table_hour.click()
            element_table_cell_edit = self.driver.find_element_by_id(self.somax_label_scheduling_table_edit_input_id)

            element_table_cell_edit.clear()
            time.sleep(0.5)
            element_table_cell_edit.send_keys(hours)
            time.sleep(0.5)
            # element_table_cell_edit.send_keys(Keys.ENTER)
            somax_logo_id_element = self.driver.find_element_by_id(self.somax_logo_id)
            # somax_logo_id_element.click()
            self.driver.execute_script("arguments[0].click();", somax_logo_id_element)

            self.wait_loading(self.somax_label_scheduling_loading_grey_table_id, timeout=3)

            # time.sleep(5)
            # print(element_table_row)
            # print(element_table_row.find_elements_by_tag_name('td')[1].text)
        print('finish-WO:{work_order}-hrs:{hours}'.format(work_order=work_order, hours=hours))
        # WorkScheduleReviseDAO.update_tasks_sync_to_somax(work_order, mode='yes')

    def wait_loading(self, loading_id, timeout=10):

        # if mode == 1:
        try:
            WebDriverWait(self.driver, timeout).until(
            EC.presence_of_element_located((By.ID, loading_id)))
            WebDriverWait(self.driver, timeout).until_not(
                EC.presence_of_element_located((By.ID, loading_id)))
        except Exception as e:
            print('loading time out')
        # elif mode == 2:
        #     WebDriverWait(self.driver, 60).until(
        #     EC.presence_of_element_located((By.ID, loading_id)))
        #     WebDriverWait(self.driver, 60).until_not(
        #         EC.presence_of_element_located((By.ID, loading_id)))

        return True

    def get_work_order_detail(self, work_order):

        try:
            element_text = self.driver.find_element_by_id(self.somax_task_first_row_work_order_a_id).text
        except Exception as e:
            element_text = ''

        if element_text != work_order:
            element_filter = WebDriverWait(self.driver, 60).until(
                EC.presence_of_element_located((By.ID, self.somax_task_filter_work_order_id)))
            self.driver.execute_script("arguments[0].setAttribute('value', {work_order})".format(work_order=work_order),
                                       element_filter)
            element_filter.send_keys(Keys.ENTER)

            WebDriverWait(self.driver, 60).until(
                EC.text_to_be_present_in_element((By.ID, self.somax_task_first_row_work_order_a_id), str(work_order)))

        element_click = WebDriverWait(self.driver, 60).until(
            EC.element_to_be_clickable((By.ID, self.somax_task_first_row_work_order_a_id)))
        element_click.click()

        return True

    def get_and_ready(self, url):
        self.driver.get(url)
        while True:
            if self.driver.current_url == url:
                break
            time.sleep(1)
        return True

    def download_finished(self, time_out=120, file_name=None, file_type=None):

        t0 = time.time()
        while True:

            if file_name:
                file_path = os.path.join(self.download_path, file_name)
            elif file_type:
                file_path = os.path.join(self.download_path, '*.{file_type}'.format(file_type=file_type))
            else:
                return False
            if os.path.exists(file_path):
                break

            t1 = time.time()
            if (t1 - t0) >= time_out:
                break

        return True

    @staticmethod
    def get_download_file_name(before, after, download_path):
        change = set(after) - set(before)

        if len(change) == 1:
            while True:
                if str(list(change)[0]).endswith('.csv'):
                    break
                after = os.listdir(download_path)
                change = set(after) - set(before)
                time.sleep(1)
            file_name = change.pop()
        else:
            return None

        return file_name

    def spider(self):
        # self.equipment_spider()
        # self.pm_spider()
        self.task_spider()

    @classmethod
    def to_excel(cls, row, value):
        wb = load_workbook('/home/arthurtu/projects/bbu/spider/somax/Work hour result.xlsx')
        ws = wb.get_active_sheet()
        ws.cell(row=row, column=16).value = value

        wb.save('/home/arthurtu/projects/bbu/spider/somax/Work hour result.xlsx')


if __name__ == '__main__':
    # SomaxSpider().task_edit_spider([{'work_order': '17037018',
    #                                  'current_status': 'Scheduled',
    #                                  'current_status_somax': 'Scheduled'}])
    # start = input('start:')
    # end = input('end:')
    # SomaxSpider().task_actual_hour_spider(start, end)

    # spider_type = input('please choose spider type: equip or 1 | pm or 2 | task or 3:')
    # if spider_type in ['equip', '1']:
    #     SomaxSpider().equipment_spider()
    # elif spider_type in ['pm', '2']:
    #     SomaxSpider().equipment_spider()
    # elif spider_type in ['task', '3']:
    #     SomaxSpider().task_spider()
    # SomaxSpider().sync_schedules_to_somax_spider()
    SomaxSpider().sync_schedules_to_somax_spider()
